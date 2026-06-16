"""
================================================================================
Pairs Trading Strategy on S&P 500 Stocks  --  Full Backtesting Framework
QF621 Quantitative Trading Strategies, Group 9
================================================================================

A configurable statistical-arbitrage pairs-trading engine that implements EVERY
variant listed in the project proposal (section 6), all driven by `Config`:

  6.1 Asset universe ........ S&P 500 names  OR  liquid ETFs
  6.2 Pair selection ........ cointegration (Engle-Granger) | distance (Gatev
                              2006) | correlation (OLS R^2)
  6.3 Pair restrictions ..... same GICS sector | market-cap band | company age
                              | corporate-finance screen | PCA-cluster matching
  6.4 Trading parameters .... N pairs, entry / exit / stop thresholds,
                              daily vs VIX-adjusted entry (trade more when calm)
  6.5 Capital allocation .... fixed equal | dynamic (concentrate when in cash)
                              | GARCH inverse-volatility | volatility targeting

Pipeline (per rolling window):
    build candidates -> apply restrictions -> rank & pick top-N -> z-score
    signals (entry/exit/stop) -> per-pair P&L -> allocate capital -> aggregate.

DATA
----
Prices: yfinance (needs internet). Point-in-time S&P 500 membership, GICS
sectors, market cap and fundamentals should come from CRSP/Compustat via WRDS
(hooks + documented schema below); a static large-cap fallback + real sector
map are provided so the framework runs out of the box.

Requires:  pip install yfinance statsmodels pandas numpy arch scikit-learn
Run:       python pairs_trading.py
================================================================================
"""

from __future__ import annotations

import warnings
from dataclasses import dataclass, field, replace
from itertools import combinations

import numpy as np
import pandas as pd
import statsmodels.api as sm
from statsmodels.tsa.adfvalues import mackinnonp
from statsmodels.tsa.stattools import coint

warnings.filterwarnings("ignore")


# =============================================================================
# Configuration -- every proposal parameter is a toggle here
# =============================================================================
@dataclass
class Config:
    # --- 6.1 universe & sample ---
    universe: str = "sp500"          # "sp500" | "etf"
    tickers: list = field(default_factory=list)   # override universe if given
    start: str = "2015-01-01"
    end: str = "2019-12-31"

    # --- rolling window (trading days) ---
    formation_days: int = 252        # ~12 months (pair selection)
    trading_days: int = 126          # ~6  months (trading)
    step_days: int = 126             # re-estimate every ~6 months

    # --- 6.2 pair-selection method ---
    method: str = "cointegration"    # "cointegration" | "distance" | "correlation"
    p_value_threshold: float = 0.05  # cointegration significance
    cointegration_robustness: bool = False
    coint_maxlag: int = 1
    coint_autolag: str = None
    fast_cointegration: bool = True
    test_both_directions: bool = False
    recent_p_value_threshold: float = 0.10
    recent_fraction: float = 0.50    # confirm coint on latest formation data
    min_r2: float = 0.80             # correlation-method minimum R^2
    correlation_on_returns: bool = False
    correlation_spread_mode: str = "ols"  # "ols" | "log_ratio" | "normalized"
    min_recent_corr: float = 0.0
    recent_corr_days: int = 63
    require_spread_reversion: bool = False
    max_beta_drift: float = np.inf   # relative first-half vs second-half beta
    validation_fraction: float = 0.0 # 0 disables formation holdout validation
    validation_pool_size: int = 200  # validate only strongest raw candidates
    min_validation_return: float = -np.inf
    min_validation_trades: int = 0
    rank_by_validation: bool = False
    n_pairs: int = 10                # top-N pairs
    max_pairs_per_asset: int = 0     # 0 = unlimited
    max_pairs_per_sector: int = 0    # 0 = unlimited
    min_corr: float = 0.50           # cheap pre-filter before the slow coint test
    require_positive_beta: bool = True
    use_log_prices: bool = False     # optional proportional spread for equities
    min_half_life: float = 2.0
    max_half_life: float = 60.0
    min_mean_crossings: int = 4
    n_jobs: int = 1                  # -1 = all cores for pair tests

    # --- 6.3 pair restrictions (each independently toggleable) ---
    restrict_same_sector: bool = False
    restrict_same_industry: bool = False
    restrict_mcap: bool = False
    mcap_log_tol: float = 0.75       # |ln(mcap_a) - ln(mcap_b)| <= tol
    restrict_age: bool = False
    min_age_years: float = 5.0
    restrict_fundamentals: bool = False
    restrict_pca_cluster: bool = False
    pca_components: int = 5
    pca_clusters: int = 8

    # --- 6.4 trading rules (z-score / std-dev units) ---
    entry_z: float = 2.0
    exit_z: float = 0.0
    stop_z: float = 3.0
    reentry_z: float = 1.0           # re-arm after a stopped spread normalises
    rearm_after_stop: bool = False
    max_holding_days: int = 0        # 0 disables time exit
    zscore_lookback: int = 0         # 0 = fixed formation mean/std
    vix_adjust: bool = False         # scale entry threshold by VIX regime
    vix_scale_lo: float = 0.6        # calm  -> lower threshold -> trade more
    vix_scale_hi: float = 1.4        # turbulent -> higher threshold
    max_vix_ratio: float = np.inf    # block entries above VIX/form median
    position_sizing: str = "unit"    # "unit" | "zscore"
    max_position_scale: float = 1.0  # cap for zscore sizing

    # --- 6.5 capital allocation ---
    allocation: str = "equal"        # "equal" | "dynamic" | "garch"
    vol_target_ann: float = 0.0      # 0 disables trailing volatility targeting
    vol_target_lookback: int = 60
    vol_target_min_periods: int = 20
    vol_target_max_scale: float = np.inf

    # --- costs & accounting ---
    tc_bps: float = 10.0             # per leg, per trade (round-trip = 2 legs)
    gross_leverage: float = 1.0
    short_borrow_bps_ann: float = 0.0
    financing_bps_ann: float = 0.0   # charged on gross exposure above 1x
    ann_factor: int = 252


# =============================================================================
# Universes & metadata
#   Sectors below are REAL GICS groupings (reliable). Market cap / fundamentals
#   are illustrative placeholders -- REPLACE with CRSP/Compustat (WRDS) for the
#   real study. PCA clustering and age-from-history need no external data.
# =============================================================================
SP500_FALLBACK = [
    "AAPL", "MSFT", "GOOGL", "AMZN", "META", "NVDA",
    "JPM", "BAC", "WFC", "C", "GS", "MS", "AXP",
    "XOM", "CVX", "COP", "SLB",
    "JNJ", "PFE", "MRK", "ABBV", "UNH",
    "PG", "KO", "PEP", "WMT", "COST",
    "HD", "LOW", "MCD", "NKE",
    "T", "VZ", "DIS", "CMCSA",
    "BA", "CAT", "GE", "HON", "MMM",
    "INTC", "CSCO", "ORCL", "IBM", "QCOM", "TXN", "V", "MA",
]

ETF_UNIVERSE = [
    "SPY", "IVV", "VOO", "QQQ", "DIA", "IWM",            # broad index
    "XLF", "XLE", "XLK", "XLV", "XLI", "XLP",            # SPDR sectors
    "XLY", "XLU", "XLB", "XLRE", "XLC",
    "XOP", "OIH", "KRE", "SMH", "SOXX",                 # industry
    "GLD", "SLV", "GDX",                                # metals
    "TLT", "IEF", "SHY", "LQD", "HYG",                  # bonds
    "EEM", "EFA", "VWO", "FXI",                         # international
]

GICS_SECTOR = {
    "AAPL": "InfoTech", "MSFT": "InfoTech", "NVDA": "InfoTech", "INTC": "InfoTech",
    "CSCO": "InfoTech", "ORCL": "InfoTech", "IBM": "InfoTech", "QCOM": "InfoTech",
    "TXN": "InfoTech", "V": "InfoTech", "MA": "InfoTech",
    "GOOGL": "CommServices", "META": "CommServices", "DIS": "CommServices",
    "CMCSA": "CommServices", "T": "CommServices", "VZ": "CommServices",
    "AMZN": "ConsDiscretionary", "HD": "ConsDiscretionary", "LOW": "ConsDiscretionary",
    "MCD": "ConsDiscretionary", "NKE": "ConsDiscretionary",
    "JPM": "Financials", "BAC": "Financials", "WFC": "Financials", "C": "Financials",
    "GS": "Financials", "MS": "Financials", "AXP": "Financials",
    "XOM": "Energy", "CVX": "Energy", "COP": "Energy", "SLB": "Energy",
    "JNJ": "HealthCare", "PFE": "HealthCare", "MRK": "HealthCare",
    "ABBV": "HealthCare", "UNH": "HealthCare",
    "PG": "ConsStaples", "KO": "ConsStaples", "PEP": "ConsStaples",
    "WMT": "ConsStaples", "COST": "ConsStaples",
    "BA": "Industrials", "CAT": "Industrials", "GE": "Industrials",
    "HON": "Industrials", "MMM": "Industrials",
}


@dataclass
class Metadata:
    """
    Optional firm characteristics for the 6.3 restrictions. Defaults to the
    static maps above; supply your own (from WRDS) to run the real study.

      sectors      : {ticker -> GICS sector str}
      mcap         : {ticker -> market cap}            (for the mcap band)
      ipo_year     : {ticker -> first-listing year}    (for the age filter)
      fundamentals : DataFrame index=ticker, columns include
                     ['profitability', 'gross_margin', 'cfo']  (corporate screen)
    """
    sectors: dict = field(default_factory=lambda: dict(GICS_SECTOR))
    industries: dict = field(default_factory=dict)
    mcap: dict = field(default_factory=dict)
    ipo_year: dict = field(default_factory=dict)
    fundamentals: pd.DataFrame = field(default_factory=pd.DataFrame)
    membership: pd.DataFrame = field(default_factory=pd.DataFrame)

    def members_at(self, date) -> set:
        """Point-in-time members at `date`; empty means no membership filter."""
        if self.membership.empty:
            return set()
        m = self.membership.copy()
        if "ticker" not in m.columns:
            m = m.reset_index().rename(columns={m.index.name or "index": "ticker"})
        start = pd.to_datetime(m["start"], errors="coerce")
        end = pd.to_datetime(m["end"], errors="coerce").fillna(pd.Timestamp.max)
        mask = start.le(pd.Timestamp(date)) & end.ge(pd.Timestamp(date))
        return set(m.loc[mask, "ticker"].astype(str))


def load_universe(name: str) -> list:
    """
    Quick-test universes. For the real S&P 500 study, pull point-in-time
    membership from CRSP/WRDS  (crsp_a_indexes.dsp500list joined to
    crsp.stocknames) so each window only sees names alive on those dates.
    """
    return {"sp500": SP500_FALLBACK, "etf": ETF_UNIVERSE}[name]


# =============================================================================
# Data layer
# =============================================================================
def download_prices(tickers: list, start: str, end: str) -> pd.DataFrame:
    """Daily adjusted-close prices via yfinance. Requires internet access."""
    import yfinance as yf
    raw = yf.download(tickers, start=start, end=end,
                      auto_adjust=True, progress=False, group_by="column")
    px = raw["Close"] if isinstance(raw.columns, pd.MultiIndex) else raw
    px = px.dropna(axis=1, how="all").sort_index()
    keep = px.columns[px.notna().mean() > 0.95]
    return px[keep].ffill().dropna(how="any")


def download_vix(start: str, end: str) -> pd.Series:
    """CBOE VIX (^VIX) close, for the 6.4 VIX-adjusted entry threshold."""
    import yfinance as yf
    raw = yf.download("^VIX", start=start, end=end, auto_adjust=False, progress=False)
    s = raw["Close"]
    return (s.iloc[:, 0] if isinstance(s, pd.DataFrame) else s).rename("VIX")


# =============================================================================
# 6.3 Pair restrictions -> candidate generation
# =============================================================================
def _pca_clusters(form_prices: pd.DataFrame, cfg: Config) -> dict:
    """Project names onto top PCA factors of returns, then KMeans-cluster them."""
    from sklearn.decomposition import PCA
    from sklearn.cluster import KMeans

    rets = form_prices.pct_change(fill_method=None).dropna()
    if rets.shape[0] < cfg.pca_components + 2 or rets.shape[1] < 2:
        return {c: 0 for c in form_prices.columns}
    X = ((rets - rets.mean()) / rets.std().replace(0, 1)).T.values   # names x time
    k = min(cfg.pca_components, X.shape[1] - 1, X.shape[0] - 1)
    factors = PCA(n_components=max(k, 1)).fit_transform(X)
    n_clusters = min(cfg.pca_clusters, len(form_prices.columns))
    labels = KMeans(n_clusters=n_clusters, n_init=10, random_state=0).fit_predict(factors)
    return dict(zip(form_prices.columns, labels))


def build_candidates(form_prices: pd.DataFrame, cfg: Config, meta: Metadata,
                     formation_year: int) -> list:
    """All eligible (a, b) pairs after every enabled 6.3 restriction."""
    # drop degenerate series (constant over the window -> no tradeable spread)
    names = [c for c in form_prices.columns
             if form_prices[c].notna().mean() >= 0.95
             and pd.notna(form_prices[c].iloc[0])
             and pd.notna(form_prices[c].iloc[-1])
             and form_prices[c].ffill().std() > 0]
    members = meta.members_at(form_prices.index[-1])
    if members:
        names = [c for c in names if c in members]

    # company-age screen: enough listing history before the formation window
    if cfg.restrict_age:
        names = [t for t in names
                 if (formation_year - meta.ipo_year.get(t, formation_year)) >= cfg.min_age_years]

    # corporate-finance screen: keep firms passing simple quality thresholds
    if cfg.restrict_fundamentals and not meta.fundamentals.empty:
        f = meta.fundamentals
        ok = []
        for t in names:
            if t not in f.index:
                continue
            row = f.loc[t]
            if (row.get("profitability", 1) > 0 and
                    row.get("gross_margin", 1) > 0.20 and
                    row.get("cfo", 1) > 0):
                ok.append(t)
        names = ok

    clusters = _pca_clusters(form_prices[names], cfg) if cfg.restrict_pca_cluster else None

    out = []
    for a, b in combinations(names, 2):
        if cfg.restrict_same_sector and meta.sectors.get(a) != meta.sectors.get(b):
            continue
        if cfg.restrict_same_industry:
            ia, ib = meta.industries.get(a), meta.industries.get(b)
            if ia is None or ib is None or ia != ib:
                continue
        if cfg.restrict_mcap and meta.mcap:
            ma, mb = meta.mcap.get(a), meta.mcap.get(b)
            if ma is None or mb is None or abs(np.log(ma) - np.log(mb)) > cfg.mcap_log_tol:
                continue
        if clusters is not None and clusters.get(a) != clusters.get(b):
            continue
        out.append((a, b))
    return out


# =============================================================================
# 6.2 Pair-selection methods
# =============================================================================
def estimate_hedge_ratio(y: pd.Series, x: pd.Series):
    """Engle-Granger step 1: OLS  y = alpha + beta*x.  Returns (alpha, beta).
    has_constant='add' guarantees an intercept even if x is constant over the
    window (e.g. a halted/floored stock), so we always get two parameters."""
    X = sm.add_constant(np.asarray(x, dtype=float), has_constant="add")
    res = sm.OLS(np.asarray(y, dtype=float), X).fit()
    return float(res.params[0]), float(res.params[1])


def fast_coint(y: pd.Series, x: pd.Series, maxlag: int = 1):
    """
    Fast Engle-Granger with fixed-lag ADF.

    Matches statsmodels.coint(..., trend="c", autolag=None) without building
    statsmodels result objects for every pair.
    """
    yv = np.asarray(y, dtype=float)
    xv = np.asarray(x, dtype=float)
    if len(yv) != len(xv) or len(yv) <= maxlag + 3:
        return np.nan, np.nan
    if not (np.isfinite(yv).all() and np.isfinite(xv).all()):
        return np.nan, np.nan

    xc = xv - xv.mean()
    yc = yv - yv.mean()
    xx = float(xc @ xc)
    if xx <= 0:
        return np.nan, np.nan
    beta = float(xc @ yc) / xx
    alpha = float(yv.mean() - beta * xv.mean())
    resid = yv - alpha - beta * xv

    tss = float(yc @ yc)
    rss = float(resid @ resid)
    if tss > 0 and 1.0 - rss / tss >= 1.0 - 100 * np.sqrt(np.finfo(float).eps):
        return -np.inf, 0.0

    delta = np.diff(resid)
    dep = delta[maxlag:]
    level = resid[maxlag:-1]
    cols = [level]
    for lag in range(1, maxlag + 1):
        cols.append(delta[maxlag - lag:-lag])
    design = np.column_stack(cols)
    try:
        params, _, _, _ = np.linalg.lstsq(design, dep, rcond=None)
        errors = dep - design @ params
        dof = len(dep) - design.shape[1]
        if dof <= 0:
            return np.nan, np.nan
        sigma2 = float(errors @ errors) / dof
        cov = sigma2 * np.linalg.pinv(design.T @ design)
        se = float(np.sqrt(max(cov[0, 0], 0.0)))
        stat = float(params[0] / se) if se > 0 else -np.inf
    except np.linalg.LinAlgError:
        return np.nan, np.nan
    return stat, float(mackinnonp(stat, regression="c", N=2))


def cointegration_test(y: pd.Series, x: pd.Series, cfg: Config):
    if cfg.fast_cointegration and cfg.coint_autolag is None:
        return fast_coint(y, x, maxlag=cfg.coint_maxlag)
    stat, pvalue, _ = coint(
        y, x, maxlag=cfg.coint_maxlag, autolag=cfg.coint_autolag)
    return float(stat), float(pvalue)


def spread_diagnostics(spread: pd.Series) -> dict:
    """Formation-only mean-reversion diagnostics for pair selection."""
    s = spread.dropna().astype(float)
    if len(s) < 20 or s.std() == 0:
        return {"half_life": np.inf, "mean_crossings": 0}

    values = s.to_numpy(dtype=float)
    lag = values[:-1]
    delta = np.diff(values)
    lag_c = lag - lag.mean()
    denom = float(lag_c @ lag_c)
    gamma = (float(lag_c @ (delta - delta.mean())) / denom
             if denom > 0 else np.nan)
    half_life = (-np.log(2) / gamma
                 if np.isfinite(gamma) and gamma < 0 else np.inf)

    centred = s - s.mean()
    signs = np.sign(centred).replace(0, np.nan).ffill().bfill()
    crossings = int((signs != signs.shift(1)).sum() - 1)
    return {"half_life": float(half_life), "mean_crossings": max(crossings, 0)}


def select_pairs(form: pd.DataFrame, candidates: list, cfg: Config,
                 meta: Metadata = None) -> list:
    """
    Rank candidate pairs by the chosen method and return the top-N as records.
    Each record carries the info needed to form its trading spread later:
      weight_mode 'beta_shares' (cointegration/correlation): hold 1 share A vs
                  beta shares B; spread = P_A - beta*P_B.
      weight_mode 'dollar_equal' (distance): hold $1 long / $1 short on prices
                  normalised to the formation start; spread = nA - nB.
    """
    if cfg.method == "cointegration" and cfg.n_jobs != 1:
        from joblib import Parallel, delayed

        def evaluate(pair):
            local_cfg = replace(cfg, n_jobs=1, n_pairs=1)
            result = select_pairs(form, [pair], local_cfg, meta)
            return result[0] if result else None

        tested = Parallel(n_jobs=cfg.n_jobs, prefer="threads")(
            delayed(evaluate)(pair) for pair in candidates)
        recs = [rec for rec in tested if rec is not None]
        recs.sort(key=lambda r: r["score"])
        return apply_pair_caps(recs, cfg, meta)

    recs = []
    for a, b in candidates:
        if cfg.method == "cointegration":
            sa = np.log(form[a]) if cfg.use_log_prices else form[a]
            sb = np.log(form[b]) if cfg.use_log_prices else form[b]
            if abs(float(sa.corr(sb))) < cfg.min_corr:
                continue

            try:
                _, p_ab = cointegration_test(sa, sb, cfg)
                p_ba = (cointegration_test(sb, sa, cfg)[1]
                        if cfg.test_both_directions else np.inf)
            except Exception:
                continue
            if p_ba < p_ab:
                a, b, sa, sb, pval = b, a, sb, sa, p_ba
            else:
                pval = p_ab
            if pval >= cfg.p_value_threshold:
                continue
            alpha, beta = estimate_hedge_ratio(sa, sb)
            if cfg.require_positive_beta and beta <= 0:
                continue

            spread = sa - alpha - beta * sb
            diag = spread_diagnostics(spread)
            recent_pval = np.nan
            if cfg.cointegration_robustness:
                recent_n = min(len(form), max(60, int(len(form) * cfg.recent_fraction)))
                try:
                    _, recent_pval = cointegration_test(
                        sa.iloc[-recent_n:], sb.iloc[-recent_n:], cfg)
                except Exception:
                    continue
                if recent_pval >= cfg.recent_p_value_threshold:
                    continue
                if not (cfg.min_half_life <= diag["half_life"] <= cfg.max_half_life):
                    continue
                if diag["mean_crossings"] < cfg.min_mean_crossings:
                    continue

            quality = (pval * (1.0 + diag["half_life"] / cfg.max_half_life)
                       if cfg.cointegration_robustness else pval)
            recs.append({"a": a, "b": b, "score": quality, "pvalue": pval,
                         "recent_pvalue": float(recent_pval), "beta": beta,
                         **diag,
                         "weight_mode": ("log_beta" if cfg.use_log_prices
                                         else "beta_shares"),
                         "metric": "pvalue"})

        elif cfg.method == "correlation":
            sa = np.log(form[a]) if cfg.use_log_prices else form[a]
            sb = np.log(form[b]) if cfg.use_log_prices else form[b]
            if cfg.correlation_on_returns:
                ca = np.log(form[a]).diff() if cfg.use_log_prices else form[a].pct_change()
                cb = np.log(form[b]).diff() if cfg.use_log_prices else form[b].pct_change()
            else:
                ca, cb = sa, sb
            corr = float(ca.corr(cb))
            r2 = corr ** 2
            if r2 < cfg.min_r2:
                continue
            recent_corr = float(ca.iloc[-cfg.recent_corr_days:].corr(
                cb.iloc[-cfg.recent_corr_days:]))
            if not np.isfinite(recent_corr) or recent_corr < cfg.min_recent_corr:
                continue

            if cfg.correlation_spread_mode == "ols":
                alpha, beta = estimate_hedge_ratio(sa, sb)
                if cfg.require_positive_beta and beta <= 0:
                    continue
                mid = len(form) // 2
                _, beta_1 = estimate_hedge_ratio(sa.iloc[:mid], sb.iloc[:mid])
                _, beta_2 = estimate_hedge_ratio(sa.iloc[mid:], sb.iloc[mid:])
                beta_drift = abs(beta_2 - beta_1) / max(abs(beta_1), 1e-12)
                rec_weights = {
                    "beta": beta,
                    "weight_mode": ("log_beta" if cfg.use_log_prices
                                    else "beta_shares"),
                }
            elif cfg.correlation_spread_mode == "log_ratio":
                beta, beta_drift = 1.0, 0.0
                rec_weights = {"beta": beta, "weight_mode": "log_ratio"}
            elif cfg.correlation_spread_mode == "normalized":
                beta, beta_drift = 1.0, 0.0
                rec_weights = {
                    "beta": beta,
                    "norm_a": float(form[a].iloc[0]),
                    "norm_b": float(form[b].iloc[0]),
                    "weight_mode": "dollar_equal",
                }
            else:
                raise ValueError(
                    f"unknown correlation_spread_mode: {cfg.correlation_spread_mode}")
            if beta_drift > cfg.max_beta_drift:
                continue

            spread = make_spread(form, {"a": a, "b": b, **rec_weights})
            diag = spread_diagnostics(spread)
            if cfg.require_spread_reversion:
                if not (cfg.min_half_life <= diag["half_life"] <= cfg.max_half_life):
                    continue
                if diag["mean_crossings"] < cfg.min_mean_crossings:
                    continue

            validation = {"return": np.nan, "sharpe": np.nan, "trades": 0}
            if cfg.validation_fraction > 0:
                validation = formation_validation(form, a, b, cfg)
                if (not np.isfinite(validation["return"]) or
                        validation["return"] < cfg.min_validation_return or
                        validation["trades"] < cfg.min_validation_trades):
                    continue
            score = (-validation["return"] if cfg.rank_by_validation
                     and np.isfinite(validation["return"]) else -r2)
            recs.append({"a": a, "b": b, "score": score,
                         "corr": corr, "recent_corr": recent_corr,
                         "beta_drift": float(beta_drift), **diag,
                         "validation_return": validation["return"],
                         "validation_sharpe": validation["sharpe"],
                         "validation_trades": validation["trades"],
                         **rec_weights,
                         "metric": "R2"})

        elif cfg.method == "distance":
            na, nb = form[a] / form[a].iloc[0], form[b] / form[b].iloc[0]
            spread = na - nb
            ssd = float((spread ** 2).sum())
            diag = spread_diagnostics(spread)
            if cfg.require_spread_reversion:
                if not (cfg.min_half_life <= diag["half_life"] <= cfg.max_half_life):
                    continue
                if diag["mean_crossings"] < cfg.min_mean_crossings:
                    continue
            recs.append({"a": a, "b": b, "score": ssd,
                         "norm_a": float(form[a].iloc[0]), "norm_b": float(form[b].iloc[0]),
                         **diag,
                         "weight_mode": "dollar_equal", "metric": "SSD"})
        else:
            raise ValueError(f"unknown method: {cfg.method}")

    recs.sort(key=lambda r: r["score"])      # lower score = better, for all methods
    if cfg.validation_fraction > 0 and cfg.method != "correlation":
        validated = []
        pool = recs[:cfg.validation_pool_size]
        for rec in pool:
            validation = formation_validation(
                form, rec["a"], rec["b"], cfg)
            if (not np.isfinite(validation["return"]) or
                    validation["return"] < cfg.min_validation_return or
                    validation["trades"] < cfg.min_validation_trades):
                continue
            rec = {
                **rec,
                "validation_return": validation["return"],
                "validation_sharpe": validation["sharpe"],
                "validation_trades": validation["trades"],
            }
            if cfg.rank_by_validation:
                rec["score"] = -validation["return"]
            validated.append(rec)
        recs = sorted(validated, key=lambda r: r["score"])
    return apply_pair_caps(recs, cfg, meta)


def apply_pair_caps(recs: list, cfg: Config, meta: Metadata = None) -> list:
    """Take top-N records while enforcing optional asset/sector concentration."""
    if cfg.max_pairs_per_asset <= 0 and cfg.max_pairs_per_sector <= 0:
        return recs[: cfg.n_pairs]

    meta = meta or Metadata()
    selected, counts, sector_counts = [], {}, {}
    for rec in recs:
        a, b = rec["a"], rec["b"]
        if cfg.max_pairs_per_asset > 0 and (
                counts.get(a, 0) >= cfg.max_pairs_per_asset or
                counts.get(b, 0) >= cfg.max_pairs_per_asset):
            continue
        sectors = {meta.sectors.get(a), meta.sectors.get(b)} - {None}
        if cfg.max_pairs_per_sector > 0 and any(
                sector_counts.get(sec, 0) >= cfg.max_pairs_per_sector
                for sec in sectors):
            continue
        selected.append(rec)
        counts[a] = counts.get(a, 0) + 1
        counts[b] = counts.get(b, 0) + 1
        for sec in sectors:
            sector_counts[sec] = sector_counts.get(sec, 0) + 1
        if len(selected) >= cfg.n_pairs:
            break
    return selected


# =============================================================================
# 3. Spread, signals, single-pair backtest
# =============================================================================
def make_spread(prices: pd.DataFrame, rec: dict) -> pd.Series:
    if rec["weight_mode"] == "beta_shares":
        return prices[rec["a"]] - rec["beta"] * prices[rec["b"]]
    if rec["weight_mode"] == "log_beta":
        return np.log(prices[rec["a"]]) - rec["beta"] * np.log(prices[rec["b"]])
    if rec["weight_mode"] == "log_ratio":
        return np.log(prices[rec["a"]]) - np.log(prices[rec["b"]])
    return prices[rec["a"]] / rec["norm_a"] - prices[rec["b"]] / rec["norm_b"]


def leg_weights(prices: pd.DataFrame, rec: dict):
    """Prior-day dollar weights of the long/short legs (no look-ahead); gross = 1."""
    if rec["weight_mode"] == "beta_shares":
        pa = prices[rec["a"]].shift(1)
        pb = (rec["beta"] * prices[rec["b"]]).shift(1)
        d = (pa.abs() + pb.abs()).replace(0, np.nan)
        return pa / d, pb / d
    if rec["weight_mode"] == "log_beta":
        gross = 1.0 + abs(rec["beta"])
        wa = pd.Series(1.0 / gross, index=prices.index)
        wb = pd.Series(abs(rec["beta"]) / gross, index=prices.index)
        return wa, wb
    half = pd.Series(0.5, index=prices.index)
    return half, half


def generate_positions(z: pd.Series, cfg: Config, entry=None) -> pd.Series:
    """
    Stateful entry/exit/stop signal, position in {-1, 0, +1}.
      +1 long spread  (long A, short B)  when z <= -entry
      -1 short spread (short A, long B)   when z >= +entry
       0 flat after z reverts through exit_z, or after the stop is breached.
    A stopped/timed-out spread must normalise inside +/- reentry_z before
    another entry. This avoids repeated stop/re-entry churn during divergence.
    `entry` may be a per-day array (VIX-adjusted) or None (use cfg.entry_z).
    """
    zv = np.asarray(z, dtype=float)
    ev = np.full(len(zv), cfg.entry_z) if entry is None else np.asarray(entry, dtype=float)
    pos = np.zeros(len(zv))
    state = 0
    armed = True
    held = 0
    for t in range(len(zv)):
        if not np.isfinite(zv[t]) or not np.isfinite(ev[t]):
            pos[t] = state
            continue
        if state == 0:
            if not armed:
                if abs(zv[t]) <= cfg.reentry_z:
                    armed = True
                pos[t] = 0
                continue
            if zv[t] <= -ev[t]:
                state = 1
                held = 0
            elif zv[t] >= ev[t]:
                state = -1
                held = 0
        elif state == 1:
            held += 1
            stopped = (zv[t] <= -cfg.stop_z or
                       (cfg.max_holding_days > 0 and held >= cfg.max_holding_days))
            if zv[t] >= cfg.exit_z or stopped:
                state = 0
                armed = not (stopped and cfg.rearm_after_stop)
        elif state == -1:
            held += 1
            stopped = (zv[t] >= cfg.stop_z or
                       (cfg.max_holding_days > 0 and held >= cfg.max_holding_days))
            if zv[t] <= cfg.exit_z or stopped:
                state = 0
                armed = not (stopped and cfg.rearm_after_stop)
        if cfg.position_sizing == "zscore" and state != 0:
            denom = ev[t] if np.isfinite(ev[t]) and ev[t] > 0 else cfg.entry_z
            scale = min(max(abs(zv[t]) / max(denom, 1e-12), 1.0),
                        cfg.max_position_scale)
            pos[t] = state * scale
        else:
            pos[t] = state
    return pd.Series(pos, index=z.index)


def backtest_pair(seed_slice: pd.DataFrame, rec: dict, mu: float, sd: float,
                  cfg: Config, entry=None,
                  formation_spread: pd.Series = None) -> pd.DataFrame:
    """
    Per-pair daily results. `seed_slice` = one seed day + the trading window
    (seed seeds pct_change and the first position; caller drops the seed row).
    Returns columns: move (one-unit long-spread return), pos, and dz.
    Portfolio allocation and transaction costs are handled jointly later so
    changing the number of active pairs is accounted for without look-ahead.
    """
    spread = make_spread(seed_slice, rec)
    if cfg.zscore_lookback > 1 and formation_spread is not None:
        history = pd.concat([formation_spread.iloc[:-1], spread])
        rolling_mu = history.rolling(cfg.zscore_lookback,
                                     min_periods=cfg.zscore_lookback).mean()
        rolling_sd = history.rolling(cfg.zscore_lookback,
                                     min_periods=cfg.zscore_lookback).std()
        z = ((history - rolling_mu) / rolling_sd.replace(0, np.nan)).reindex(
            spread.index)
    else:
        z = (spread - mu) / sd
    pos = generate_positions(z, cfg, entry)

    rA = seed_slice[rec["a"]].pct_change(fill_method=None)
    rB = seed_slice[rec["b"]].pct_change(fill_method=None)
    wA, wB = leg_weights(seed_slice, rec)

    return pd.DataFrame({
        "move": (wA * rA - wB * rB).fillna(0.0),
        "pos": pos,
        "dz": spread.diff() / sd,
    })


# =============================================================================
# 6.5 Capital allocation
# =============================================================================
def _garch_vol(recs: list, form: pd.DataFrame, dz_df: pd.DataFrame) -> pd.DataFrame:
    """
    GARCH(1,1) one-step-ahead volatility forecast per pair, on the z-scored
    spread change (dimensionless -> comparable across pairs). Fit params on the
    formation window, then roll the conditional-variance recursion forward over
    the trading window using realised dz (info up to t-1 only).
    """
    from arch import arch_model
    cols = {}
    for rec in recs:
        key = f'{rec["a"]}-{rec["b"]}'
        if key not in dz_df:
            continue
        sd = make_spread(form, rec).std()
        dz_form = (make_spread(form, rec).diff().dropna() / sd).values
        dz_form = dz_form - dz_form.mean()
        try:
            fit = arch_model(dz_form, mean="Zero", vol="GARCH", p=1, q=1,
                             rescale=False).fit(disp="off")
            w_ = float(fit.params["omega"])
            a_ = float(fit.params["alpha[1]"])
            b_ = float(fit.params["beta[1]"])
            prev_var = float(np.var(dz_form))
            prev_e2 = float(dz_form[-1] ** 2)
            d = dz_df[key].fillna(0.0).values
            sig = np.empty(len(d))
            for t in range(len(d)):
                v = w_ + a_ * prev_e2 + b_ * prev_var      # forecast for day t
                sig[t] = np.sqrt(max(v, 1e-12))
                prev_var, prev_e2 = v, d[t] ** 2
            cols[key] = pd.Series(sig, index=dz_df.index)
        except Exception:
            cols[key] = dz_df[key].rolling(20, min_periods=5).std().bfill()
    return pd.DataFrame(cols)


def allocate(move_df: pd.DataFrame, pos_df: pd.DataFrame, dz_df: pd.DataFrame,
             cfg: Config, recs: list, form: pd.DataFrame) -> pd.Series:
    """
    Build close-to-close portfolio returns from target pair exposures.

    Signals observed at close t set target exposure for t+1. Gross return on
    day t therefore uses target exposure from t-1. Costs are charged on every
    target change, including initial entry, dynamic reallocation, and forced
    liquidation at the end of each trading window.
    """
    if move_df.empty:
        return pd.Series(dtype=float)

    active = pos_df.abs() > 0
    if cfg.allocation == "equal":                     # fixed 1/N (cash sits idle)
        weights = active.astype(float) / len(recs)
    elif cfg.allocation == "dynamic":                 # equal among active pairs
        n = active.sum(axis=1).replace(0, np.nan)
        weights = active.div(n, axis=0).fillna(0.0)
    elif cfg.allocation == "garch":                   # inverse-vol among active pairs
        vol = _garch_vol(recs, form, dz_df).reindex_like(move_df)
        inv = (1.0 / vol).replace([np.inf, -np.inf], np.nan)
        raw = inv.where(active, 0.0).fillna(0.0)
        s = raw.sum(axis=1).replace(0, np.nan)
        weights = raw.div(s, axis=0).fillna(0.0)
    else:
        raise ValueError(f"unknown allocation: {cfg.allocation}")

    target = weights * pos_df * cfg.gross_leverage
    if cfg.vol_target_ann > 0:
        proxy = (target.shift(1).fillna(0.0) * move_df.fillna(0.0)).sum(axis=1)
        realised = (
            proxy.shift(1)
            .rolling(cfg.vol_target_lookback,
                     min_periods=cfg.vol_target_min_periods)
            .std()
            * np.sqrt(cfg.ann_factor)
        )
        scale = (cfg.vol_target_ann / realised).replace([np.inf, -np.inf], np.nan)
        scale = scale.clip(lower=0.0, upper=cfg.vol_target_max_scale).fillna(1.0)
        target = target.mul(scale, axis=0)

    gross = (target.shift(1).fillna(0.0) * move_df.fillna(0.0)).sum(axis=1)
    turnover = target.diff().abs().sum(axis=1)
    turnover.iloc[0] = target.iloc[0].abs().sum()
    prior_gross = target.shift(1).fillna(0.0).abs().sum(axis=1)
    borrow = (0.5 * prior_gross * cfg.short_borrow_bps_ann /
              1e4 / cfg.ann_factor)
    financed = (prior_gross - 1.0).clip(lower=0.0)
    financing = financed * cfg.financing_bps_ann / 1e4 / cfg.ann_factor
    net = gross - turnover * (cfg.tc_bps / 1e4) - borrow - financing

    # The seed row represents the formation-window close. Carry its setup cost
    # into the first reported trading day, then liquidate at the window end.
    if len(net) > 1:
        net.iloc[1] += net.iloc[0]
        net = net.iloc[1:].copy()
        net.iloc[-1] -= target.iloc[-1].abs().sum() * (cfg.tc_bps / 1e4)
    else:
        net = net.iloc[0:0]
    return net


def formation_validation(form: pd.DataFrame, a: str, b: str,
                         cfg: Config) -> dict:
    """Train on early formation data and validate the pair on its later part."""
    split = int(len(form) * (1.0 - cfg.validation_fraction))
    if split < 60 or len(form) - split < 40:
        return {"return": np.nan, "sharpe": np.nan, "trades": 0}

    train = form.iloc[:split]
    seed_validation = form.iloc[split - 1:]
    if cfg.method == "distance":
        rec = {
            "a": a, "b": b,
            "norm_a": float(train[a].iloc[0]),
            "norm_b": float(train[b].iloc[0]),
            "weight_mode": "dollar_equal",
        }
    elif cfg.method == "cointegration":
        sa = np.log(train[a]) if cfg.use_log_prices else train[a]
        sb = np.log(train[b]) if cfg.use_log_prices else train[b]
        _, beta = estimate_hedge_ratio(sa, sb)
        if cfg.require_positive_beta and beta <= 0:
            return {"return": np.nan, "sharpe": np.nan, "trades": 0}
        rec = {
            "a": a, "b": b, "beta": beta,
            "weight_mode": "log_beta" if cfg.use_log_prices else "beta_shares",
        }
    elif cfg.correlation_spread_mode == "ols":
        sa = np.log(train[a]) if cfg.use_log_prices else train[a]
        sb = np.log(train[b]) if cfg.use_log_prices else train[b]
        _, beta = estimate_hedge_ratio(sa, sb)
        if cfg.require_positive_beta and beta <= 0:
            return {"return": np.nan, "sharpe": np.nan, "trades": 0}
        rec = {
            "a": a, "b": b, "beta": beta,
            "weight_mode": "log_beta" if cfg.use_log_prices else "beta_shares",
        }
    elif cfg.correlation_spread_mode == "log_ratio":
        rec = {"a": a, "b": b, "beta": 1.0, "weight_mode": "log_ratio"}
    elif cfg.correlation_spread_mode == "normalized":
        rec = {
            "a": a, "b": b, "beta": 1.0,
            "norm_a": float(train[a].iloc[0]),
            "norm_b": float(train[b].iloc[0]),
            "weight_mode": "dollar_equal",
        }
    else:
        return {"return": np.nan, "sharpe": np.nan, "trades": 0}
    formation_spread = make_spread(train, rec)
    sd = formation_spread.std()
    if not np.isfinite(sd) or sd == 0:
        return {"return": np.nan, "sharpe": np.nan, "trades": 0}

    bt = backtest_pair(seed_validation, rec, formation_spread.mean(), sd, cfg,
                       formation_spread=formation_spread)
    key = f"{a}-{b}"
    move_df = pd.DataFrame({key: bt["move"]})
    pos_df = pd.DataFrame({key: bt["pos"]})
    dz_df = pd.DataFrame({key: bt["dz"]})
    validation_cfg = replace(cfg, allocation="equal")
    returns = allocate(move_df, pos_df, dz_df, validation_cfg, [rec], train)
    entries = int(((bt["pos"] != 0) & (bt["pos"].shift(1).fillna(0) == 0)).sum())
    if returns.empty:
        return {"return": np.nan, "sharpe": np.nan, "trades": entries}
    total = float((1.0 + returns).prod() - 1.0)
    vol = float(returns.std())
    sharpe = (float(returns.mean()) / vol * np.sqrt(cfg.ann_factor)
              if vol > 0 else np.nan)
    return {"return": total, "sharpe": sharpe, "trades": entries}


# =============================================================================
# 4. Rolling-window backtest (orchestration)
# =============================================================================
def _entry_series(seed_idx, form_idx, cfg: Config, vix: pd.Series):
    """Per-day entry threshold over the seed slice (VIX-adjusted if enabled)."""
    if vix is None or (not cfg.vix_adjust and not np.isfinite(cfg.max_vix_ratio)):
        return None
    ref = vix.reindex(form_idx).median()
    if not np.isfinite(ref) or ref == 0:
        return None
    ratio = vix.reindex(seed_idx) / ref
    if cfg.vix_adjust:
        scale = ratio.clip(cfg.vix_scale_lo, cfg.vix_scale_hi)
        entry = (cfg.entry_z * scale).fillna(cfg.entry_z)
    else:
        entry = pd.Series(cfg.entry_z, index=seed_idx, dtype=float)
    if np.isfinite(cfg.max_vix_ratio):
        entry = entry.mask(ratio > cfg.max_vix_ratio, np.inf)
    return entry.values


def run_strategy(prices: pd.DataFrame, cfg: Config,
                 meta: Metadata = None, vix: pd.Series = None) -> dict:
    meta = meta or Metadata()
    idx = prices.index
    n = len(idx)
    port = pd.Series(0.0, index=idx)
    selections = []

    start = 0
    while start + cfg.formation_days + cfg.trading_days <= n:
        f0, f1 = start, start + cfg.formation_days
        t1 = f1 + cfg.trading_days
        form = prices.iloc[f0:f1]
        f_year = idx[f1 - 1].year

        candidates = build_candidates(form, cfg, meta, f_year)
        recs = select_pairs(form, candidates, cfg, meta)
        selections.append({"formation": (idx[f0], idx[f1 - 1]),
                           "trading": (idx[f1], idx[t1 - 1]),
                           "n_candidates": len(candidates),
                           "n_selected": len(recs),
                           "pairs": [(r["a"], r["b"]) for r in recs]})

        if recs:
            seed = prices.iloc[f1 - 1:t1]
            entry = _entry_series(seed.index, form.index, cfg, vix)
            moves, poss, dzs, kept = {}, {}, {}, []
            for rec in recs:
                fs = make_spread(form, rec)
                mu, sd = fs.mean(), fs.std()
                if sd == 0 or np.isnan(sd):
                    continue
                bt = backtest_pair(seed, rec, mu, sd, cfg, entry, fs)
                key = f'{rec["a"]}-{rec["b"]}'
                moves[key], poss[key], dzs[key] = bt["move"], bt["pos"], bt["dz"]
                kept.append(rec)
            if kept:
                move_df = pd.DataFrame(moves)
                pos_df = pd.DataFrame(poss)
                dz_df = pd.DataFrame(dzs)
                w = allocate(move_df, pos_df, dz_df, cfg, kept, form)
                port.loc[w.index] = w.values

        start += cfg.step_days

    return {"returns": port, "selections": selections}


def run_ensemble(prices: pd.DataFrame, configs: dict,
                 meta: Metadata = None, vix: pd.Series = None,
                 weights: dict = None) -> dict:
    """Combine independently costed strategy models with fixed capital weights."""
    if not configs:
        return {"returns": pd.Series(0.0, index=prices.index),
                "selections": [], "model_results": {}}
    model_results = run_configs(prices, configs, meta, vix)
    if weights is None:
        weights = {name: 1.0 for name in configs}
    total = sum(max(float(weights.get(name, 0.0)), 0.0) for name in configs)
    if total <= 0:
        raise ValueError("ensemble weights must contain a positive value")
    norm = {name: max(float(weights.get(name, 0.0)), 0.0) / total
            for name in configs}
    returns = sum(model_results[name]["returns"] * norm[name]
                  for name in configs)

    merged = []
    max_windows = max(len(r["selections"]) for r in model_results.values())
    for i in range(max_windows):
        available = [r["selections"][i] for r in model_results.values()
                     if i < len(r["selections"])]
        pairs = sorted({pair for selection in available
                        for pair in selection["pairs"]})
        merged.append({
            "formation": available[0]["formation"],
            "trading": available[0]["trading"],
            "n_candidates": max(s["n_candidates"] for s in available),
            "n_selected": len(pairs),
            "pairs": pairs,
        })
    return {"returns": returns, "selections": merged,
            "model_results": model_results, "model_weights": norm}


def default_ensemble_configs(base: Config) -> dict:
    """Two fixed spread constructions used by the robust default runner."""
    return {
        "ols": replace(base, correlation_spread_mode="ols"),
        "log-ratio": replace(base, correlation_spread_mode="log_ratio"),
    }


def proposal_baseline_config(**overrides) -> Config:
    """Exact primary specification from the Group 9 project proposal."""
    cfg = Config(
        universe="sp500",
        method="cointegration",
        formation_days=252,
        trading_days=126,
        step_days=126,
        p_value_threshold=0.05,
        cointegration_robustness=False,
        n_pairs=10,
        min_corr=-1.0,               # test every eligible pair
        require_positive_beta=False,
        restrict_same_sector=False,
        entry_z=2.0,
        exit_z=0.0,
        stop_z=3.0,
        rearm_after_stop=False,
        max_holding_days=0,
        allocation="equal",
        tc_bps=10.0,
        gross_leverage=1.0,
        short_borrow_bps_ann=0.0,
        financing_bps_ann=0.0,
        n_jobs=1,
    )
    return replace(cfg, **overrides)


def optimized_proposal_config(**overrides) -> Config:
    """
    Cost-aware optimized specification that keeps the proposal's core logic:
    Engle-Granger cointegration, 12m/6m rolling windows, z-score entries,
    stop losses, VIX gating, and equal capital allocation.
    """
    cfg = proposal_baseline_config(
        method="cointegration",
        restrict_same_sector=True,
        restrict_same_industry=True,
        require_positive_beta=True,
        use_log_prices=True,
        min_corr=0.50,
        p_value_threshold=0.01,
        n_pairs=3,
        entry_z=2.0,
        exit_z=0.0,
        stop_z=3.5,
        rearm_after_stop=True,
        reentry_z=1.0,
        max_holding_days=60,
        max_pairs_per_asset=2,
        max_vix_ratio=1.5,
        allocation="equal",
    )
    return replace(cfg, **overrides)


def optimized_aggressive_config(**overrides) -> Config:
    """
    Higher-deployment version of the optimized proposal strategy.

    gross_leverage=3.0 means a market-neutral book can deploy about 150% long
    and 150% short gross exposure. This is more aggressive than the conservative
    1x setting, but keeps the same pair-selection and trading rules.
    """
    cfg = optimized_proposal_config(gross_leverage=3.0)
    return replace(cfg, **overrides)


def optimized_vol_target_config(**overrides) -> Config:
    """
    Main optimized specification for the report-quality run.

    It keeps the proposal's Engle-Granger pair selection and z-score trading
    rules, then uses the section-6.5 GARCH allocation plus a trailing portfolio
    volatility target. The scale cap keeps the book from becoming an unlimited
    leverage exercise during quiet windows.
    """
    cfg = optimized_proposal_config(
        allocation="garch",
        vol_target_ann=0.15,
        vol_target_max_scale=3.0,
    )
    return replace(cfg, **overrides)


def optimized_risk_balanced_config(**overrides) -> Config:
    """
    Lower-volatility companion to the headline optimized strategy.

    The same pair selection and trading rules are used, but the volatility
    target is lower. This gives a cleaner risk-control story for presentations:
    lower drawdown, higher average Sharpe across the tested periods, and still
    positive returns in the pre-COVID, post-COVID, and recent holdout samples.
    """
    cfg = optimized_proposal_config(
        allocation="garch",
        vol_target_ann=0.08,
        vol_target_max_scale=5.0,
    )
    return replace(cfg, **overrides)


# =============================================================================
# 5. Performance metrics
# =============================================================================
def performance_metrics(rets: pd.Series, cfg: Config) -> dict:
    r = rets.fillna(0.0)
    r = r.loc[r.ne(0).cummax()]
    if r.empty:
        return {}
    eq = (1 + r).cumprod()
    yrs = len(r) / cfg.ann_factor
    cagr = eq.iloc[-1] ** (1 / yrs) - 1 if yrs > 0 else np.nan
    vol = r.std() * np.sqrt(cfg.ann_factor)
    sharpe = (r.mean() * cfg.ann_factor) / vol if vol > 0 else np.nan
    dn = r[r < 0].std() * np.sqrt(cfg.ann_factor)
    sortino = (r.mean() * cfg.ann_factor) / dn if dn > 0 else np.nan
    maxdd = (eq / eq.cummax() - 1).min()
    traded = r[r != 0]
    return {
        "Total Return": float(eq.iloc[-1] - 1), "CAGR": float(cagr),
        "Ann. Vol": float(vol), "Sharpe": float(sharpe), "Sortino": float(sortino),
        "Max Drawdown": float(maxdd),
        "Calmar": float(cagr / abs(maxdd)) if maxdd < 0 else np.nan,
        "Daily Hit Rate": float((traded > 0).mean()) if len(traded) else np.nan,
        "Trading Days": int(len(r)),
    }


# =============================================================================
# 6. Optimisation helpers
# =============================================================================
def parameter_sweep(prices, base, meta=None, vix=None,
                    entry_grid=(1.5, 2.0, 2.5), npairs_grid=(5, 10, 20),
                    stop_grid=(2.5, 3.0, 3.5)) -> pd.DataFrame:
    """6.4 grid-search of the core trading parameters, ranked by Sharpe."""
    rows = []
    for e in entry_grid:
        for npr in npairs_grid:
            for s in stop_grid:
                cfg = replace(base, entry_z=e, n_pairs=npr, stop_z=s)
                p = performance_metrics(run_strategy(prices, cfg, meta, vix)["returns"], cfg)
                if p:
                    rows.append({"entry_z": e, "n_pairs": npr, "stop_z": s,
                                 "Sharpe": p["Sharpe"], "CAGR": p["CAGR"],
                                 "MaxDD": p["Max Drawdown"]})
    return pd.DataFrame(rows).sort_values("Sharpe", ascending=False).reset_index(drop=True)


def run_configs(prices, configs: dict, meta=None, vix=None) -> dict:
    """Run several named Config variants; returns {name: run_strategy result}."""
    return {name: run_strategy(prices, cfg, meta, vix) for name, cfg in configs.items()}


def compare_configs(prices, configs: dict, meta=None, vix=None, results: dict = None):
    """Tabulate metrics for named Config variants side by side (one column each)."""
    results = results or run_configs(prices, configs, meta, vix)
    out = {name: performance_metrics(results[name]["returns"], configs[name]) for name in configs}
    return pd.DataFrame(out).round(4)


# =============================================================================
# 7. Visualisation  (matplotlib; saves PNGs you can drop straight into a report)
# =============================================================================
def _equity_drawdown(returns: pd.Series):
    r = returns.fillna(0.0)
    r = r.loc[r.ne(0).cummax()]
    eq = (1 + r).cumprod()
    return eq, eq / eq.cummax() - 1


def pair_selection_matrix(selections: list) -> pd.DataFrame:
    """Binary matrix (pair x rolling window): was the pair selected that window?"""
    labels = [f"{s['trading'][0].date()}" for s in selections]
    allp = sorted({f"{a}-{b}" for s in selections for a, b in s["pairs"]})
    mat = pd.DataFrame(0, index=allp, columns=labels)
    for lab, s in zip(labels, selections):
        for a, b in s["pairs"]:
            mat.loc[f"{a}-{b}", lab] = 1
    return mat


def plot_backtest(res: dict, title: str = "Pairs Trading Backtest", save_path: str = None):
    """Single-run overview: cumulative NAV, underwater drawdown, pairs/window."""
    import matplotlib.pyplot as plt
    eq, dd = _equity_drawdown(res["returns"])
    sel = res["selections"]
    xs = [s["trading"][0] for s in sel]
    ns = [s["n_selected"] for s in sel]

    fig, ax = plt.subplots(3, 1, figsize=(12, 11),
                           gridspec_kw={"height_ratios": [3, 2, 2]}, sharex=False)
    ax[0].plot(eq.index, eq.values, lw=1.6, color="#1f4e79")
    ax[0].axhline(1, color="grey", lw=0.8, ls="--")
    ax[0].set_ylabel("Growth of $1"); ax[0].set_title(title, fontweight="bold")
    ax[0].grid(alpha=0.3)

    ax[1].fill_between(dd.index, dd.values * 100, 0, color="#c0392b", alpha=0.5)
    ax[1].set_ylabel("Drawdown (%)"); ax[1].grid(alpha=0.3)

    ax[2].bar(xs, ns, width=80, color="#2e7d32", alpha=0.8)
    ax[2].set_ylabel("Pairs selected"); ax[2].set_xlabel("Rolling window (trading start)")
    ax[2].grid(alpha=0.3, axis="y")
    fig.tight_layout()
    if save_path:
        fig.savefig(save_path, dpi=150, bbox_inches="tight"); plt.close(fig)
    return save_path


def plot_comparison(results: dict, configs: dict, comp_df: pd.DataFrame, save_path: str = None):
    """Variant comparison: overlaid equity curves + Sharpe / CAGR / MaxDD bars."""
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(2, 2, figsize=(14, 9))
    cmap = plt.cm.tab10(np.linspace(0, 1, len(results)))

    for (name, res), c in zip(results.items(), cmap):
        eq, _ = _equity_drawdown(res["returns"])
        ax[0, 0].plot(eq.index, eq.values, lw=1.3, label=name, color=c)
    ax[0, 0].axhline(1, color="grey", lw=0.8, ls="--")
    ax[0, 0].set_title("Equity curves", fontweight="bold")
    ax[0, 0].set_ylabel("Growth of $1"); ax[0, 0].legend(fontsize=7); ax[0, 0].grid(alpha=0.3)

    names = list(comp_df.columns)
    for axi, metric, col in [(ax[0, 1], "Sharpe", "#1f4e79"),
                             (ax[1, 0], "CAGR", "#2e7d32"),
                             (ax[1, 1], "Max Drawdown", "#c0392b")]:
        vals = comp_df.loc[metric].values.astype(float)
        if metric in ("CAGR", "Max Drawdown"):
            vals = vals * 100
        axi.barh(names, vals, color=col, alpha=0.85)
        axi.set_title(metric + (" (%)" if metric != "Sharpe" else ""), fontweight="bold")
        axi.grid(alpha=0.3, axis="x"); axi.invert_yaxis()
    fig.tight_layout()
    if save_path:
        fig.savefig(save_path, dpi=150, bbox_inches="tight"); plt.close(fig)
    return save_path


def plot_pair_heatmap(selections: list, save_path: str = None):
    """Heatmap of which pairs were selected in which window (turnover / stability)."""
    import matplotlib.pyplot as plt
    mat = pair_selection_matrix(selections)
    fig, axi = plt.subplots(figsize=(max(8, 0.5 * mat.shape[1] + 3), max(4, 0.3 * mat.shape[0] + 1)))
    axi.imshow(mat.values, aspect="auto", cmap="Blues", vmin=0, vmax=1)
    axi.set_yticks(range(mat.shape[0])); axi.set_yticklabels(mat.index, fontsize=7)
    axi.set_xticks(range(mat.shape[1])); axi.set_xticklabels(mat.columns, fontsize=7, rotation=90)
    axi.set_title("Pair selection across windows", fontweight="bold")
    axi.set_xlabel("Trading-window start")
    fig.tight_layout()
    if save_path:
        fig.savefig(save_path, dpi=150, bbox_inches="tight"); plt.close(fig)
    return save_path


# =============================================================================
# 8. Excel export  (formatted results workbook -- backtest OUTPUTS, not a model)
# =============================================================================
def export_excel(path: str, comp_df: pd.DataFrame, base_res: dict,
                 sweep_df: pd.DataFrame = None):
    """Write a formatted multi-sheet results workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEAD = PatternFill("solid", fgColor="1F4E79")
    HFONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    BASE = Font(name="Arial", size=10)
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    PCT_ROWS = {"Total Return", "CAGR", "Ann. Vol", "Max Drawdown", "Daily Hit Rate"}

    wb = Workbook()

    # --- Sheet 1: variant comparison (metrics x variants) ---
    ws = wb.active; ws.title = "Variant Comparison"
    ws.cell(1, 1, "Metric").font = HFONT; ws.cell(1, 1).fill = HEAD
    for j, name in enumerate(comp_df.columns, start=2):
        c = ws.cell(1, j, name); c.font = HFONT; c.fill = HEAD; c.alignment = Alignment(horizontal="center")
    for i, metric in enumerate(comp_df.index, start=2):
        ws.cell(i, 1, metric).font = Font(name="Arial", bold=True, size=10)
        for j, name in enumerate(comp_df.columns, start=2):
            v = comp_df.loc[metric, name]
            c = ws.cell(i, j, float(v) if pd.notna(v) else None); c.font = BASE; c.border = BORDER
            c.alignment = Alignment(horizontal="center")
            if metric in PCT_ROWS:
                c.number_format = "0.0%;(0.0%);-"
            elif metric == "Trading Days":
                c.number_format = "0"
            else:
                c.number_format = "0.00;(0.00);-"
    ws.column_dimensions["A"].width = 16
    for j in range(2, comp_df.shape[1] + 2):
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.freeze_panes = "B2"

    # --- Sheet 2: selected pairs per window (base run) ---
    ws2 = wb.create_sheet("Selected Pairs")
    hdr = ["Formation start", "Formation end", "Trading start", "Trading end",
           "# candidates", "# selected", "Pairs"]
    for j, h in enumerate(hdr, start=1):
        c = ws2.cell(1, j, h); c.font = HFONT; c.fill = HEAD
    for i, s in enumerate(base_res["selections"], start=2):
        ws2.cell(i, 1, str(s["formation"][0].date())); ws2.cell(i, 2, str(s["formation"][1].date()))
        ws2.cell(i, 3, str(s["trading"][0].date())); ws2.cell(i, 4, str(s["trading"][1].date()))
        ws2.cell(i, 5, s["n_candidates"]); ws2.cell(i, 6, s["n_selected"])
        ws2.cell(i, 7, ", ".join(f"{a}-{b}" for a, b in s["pairs"]))
        for j in range(1, 8):
            ws2.cell(i, j).font = BASE
    for j, w in zip(range(1, 8), [15, 15, 14, 14, 12, 11, 70]):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.freeze_panes = "A2"

    # --- Sheet 3: pair selection frequency ---
    ws3 = wb.create_sheet("Pair Frequency")
    freq = pair_selection_matrix(base_res["selections"]).sum(axis=1).sort_values(ascending=False)
    ws3.cell(1, 1, "Pair").font = HFONT; ws3.cell(1, 1).fill = HEAD
    ws3.cell(1, 2, "Windows selected").font = HFONT; ws3.cell(1, 2).fill = HEAD
    for i, (p, n) in enumerate(freq.items(), start=2):
        ws3.cell(i, 1, p).font = BASE; ws3.cell(i, 2, int(n)).font = BASE
    ws3.column_dimensions["A"].width = 16; ws3.column_dimensions["B"].width = 18
    ws3.freeze_panes = "A2"

    # --- Sheet 4: parameter sweep (optional) ---
    if sweep_df is not None and not sweep_df.empty:
        ws4 = wb.create_sheet("Param Sweep")
        for j, h in enumerate(sweep_df.columns, start=1):
            c = ws4.cell(1, j, h); c.font = HFONT; c.fill = HEAD
        for i, (_, row) in enumerate(sweep_df.iterrows(), start=2):
            for j, h in enumerate(sweep_df.columns, start=1):
                c = ws4.cell(i, j, float(row[h])); c.font = BASE
                c.number_format = "0.0%" if h in ("CAGR", "MaxDD") else "0.00" if h == "Sharpe" else "0.0"
        for j in range(1, len(sweep_df.columns) + 1):
            ws4.column_dimensions[get_column_letter(j)].width = 12
        ws4.freeze_panes = "A2"

    wb.save(path)
    return path


# =============================================================================
# Runner
# =============================================================================
def run_period(label, start, end, base, meta=None, vix=None, verbose=True):
    cfg = replace(base, start=start, end=end,
                  tickers=(base.tickers or load_universe(base.universe)))
    prices = download_prices(cfg.tickers, start, end)
    vix_s = vix if vix is not None else (download_vix(start, end) if cfg.vix_adjust else None)
    res = run_strategy(prices, cfg, meta, vix_s)
    perf = performance_metrics(res["returns"], cfg)
    if verbose:
        avg = np.mean([s["n_selected"] for s in res["selections"]]) if res["selections"] else 0
        print(f"\n{'='*62}\n{label}: {start} -> {end}  "
              f"[{cfg.method} | alloc={cfg.allocation} | vix={cfg.vix_adjust}]\n{'='*62}")
        print(f"Universe: {prices.shape[1]} names, {prices.shape[0]} days | "
              f"windows: {len(res['selections'])} | avg pairs/window: {avg:.1f}")
        print("-" * 40)
        for k, v in perf.items():
            print(f"  {k:16s}: {v:,.4f}" if isinstance(v, float) else f"  {k:16s}: {v}")
    return {"prices": prices, "vix": vix_s, **res, "perf": perf}


if __name__ == "__main__":
    base = Config(
        universe="sp500", method="correlation", allocation="equal",
        formation_days=252, trading_days=126, step_days=126,
        n_pairs=5, entry_z=2.75, exit_z=0.25, stop_z=3.5,
        min_r2=0.25, tc_bps=10.0,
        restrict_same_sector=True,
        correlation_on_returns=True, min_recent_corr=0.30,
        max_pairs_per_asset=2, max_pairs_per_sector=3,
        require_spread_reversion=True, max_beta_drift=1.0,
        reentry_z=1.0, max_holding_days=60,
        short_borrow_bps_ann=50.0, financing_bps_ann=500.0,
    )
    meta = Metadata()   # static sector map; add mcap / fundamentals / ipo for those filters

    # --- robust two-model ensemble, both regimes ---
    periods = {
        "PRE-COVID": ("2015-01-01", "2019-12-31"),
        "POST-COVID": ("2022-01-01", "2024-12-31"),
    }
    runs = {}
    for label, (start, end) in periods.items():
        tickers = base.tickers or load_universe(base.universe)
        prices = download_prices(tickers, start, end)
        result = run_ensemble(prices, default_ensemble_configs(base), meta)
        result["prices"] = prices
        result["perf"] = performance_metrics(result["returns"], base)
        runs[label] = result
        print(f"\n{'='*62}\n{label}: {start} -> {end}  "
              f"[correlation ensemble | allocation=equal]\n{'='*62}")
        for k, v in result["perf"].items():
            print(f"  {k:16s}: {v:,.4f}" if isinstance(v, float)
                  else f"  {k:16s}: {v}")
    pre, post = runs["PRE-COVID"], runs["POST-COVID"]
    print(f"\n{'='*62}\nPRE vs POST-COVID\n{'='*62}")
    print(pd.DataFrame({"Pre-COVID": pre["perf"], "Post-COVID": post["perf"]}).round(4).to_string())

    # --- ablation across the proposal's variants (reuses pre-COVID prices) ---
    px, vx = pre["prices"], None
    variants = {
        "correlation/ols":     replace(base, correlation_spread_mode="ols"),
        "correlation/log-ratio": replace(base, correlation_spread_mode="log_ratio"),
        "coint/equal":        replace(base, method="cointegration", allocation="equal"),
        "distance/equal":     replace(base, method="distance", allocation="equal"),
        "corr/all-sectors":   replace(base, restrict_same_sector=False),
    }
    var_res = run_configs(px, variants, meta, vx)              # run each variant ONCE
    var_res = {"robust ensemble": pre, **var_res}
    variants = {"robust ensemble": base, **variants}
    comp = compare_configs(px, variants, meta, vx, results=var_res)
    print(f"\n{'='*62}\nVARIANT COMPARISON (pre-COVID)\n{'='*62}")
    print(comp.to_string())

    # --- charts + results workbook ---
    plot_backtest(pre, "Pairs Trading - Robust Ensemble (pre-COVID)", "backtest_overview.png")
    plot_comparison(var_res, variants, comp, "variant_comparison.png")
    plot_pair_heatmap(pre["selections"], "pair_selection_heatmap.png")
    # sweep = parameter_sweep(px, base, meta)                  # optional (6.4); slow
    export_excel("pairs_backtest_results.xlsx", comp, pre, sweep_df=None)
    print("\nSaved: backtest_overview.png, variant_comparison.png, "
          "pair_selection_heatmap.png, pairs_backtest_results.xlsx")
